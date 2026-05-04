"""
RR Engine — Unit Economics v1
Gera 02_unit_economics.xlsx com 4 abas:
- Premissas (todos os inputs editaveis em azul)
- Calculo (custo de tokens por agente nos 3 cenarios)
- Cenarios (margem por plano e volume)
- Stress test (sensibilidades)

Saida: 02_unit_economics.xlsx no mesmo diretorio.

Convencoes de cor (industria):
- Azul: input editavel
- Preto: formula
- Verde: link entre abas
- Amarelo (fundo): destaque
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).parent / "02_unit_economics.xlsx"

wb = Workbook()
wb.remove(wb.active)

# ===== Estilos =====
F_TITLE = Font(name='Arial', size=14, bold=True, color='000000')
F_SECTION = Font(name='Arial', size=11, bold=True, color='000000')
F_HEADER = Font(name='Arial', size=10, bold=True, color='FFFFFF')
F_INPUT = Font(name='Arial', size=10, color='0000FF')
F_INPUT_BOLD = Font(name='Arial', size=10, color='0000FF', bold=True)
F_FORMULA = Font(name='Arial', size=10, color='000000')
F_FORMULA_BOLD = Font(name='Arial', size=10, color='000000', bold=True)
F_LINK = Font(name='Arial', size=10, color='008000')
F_NOTE = Font(name='Arial', size=9, italic=True, color='666666')

FILL_HEADER = PatternFill('solid', start_color='305496')
FILL_SECTION = PatternFill('solid', start_color='D9E1F2')
FILL_HIGHLIGHT = PatternFill('solid', start_color='FFFF00')
FILL_TOTAL = PatternFill('solid', start_color='F2F2F2')

THIN = Side(border_style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

FMT_BRL = 'R$ #,##0.00;[Red](R$ #,##0.00);-'
FMT_USD = '$#,##0.0000;[Red]($#,##0.0000);-'
FMT_INT = '#,##0;[Red](#,##0);-'
FMT_PCT = '0.0%;[Red](0.0%);-'
FMT_TOK = '#,##0'

def set_cell(ws, ref, value, font=None, fill=None, fmt=None, align=None, border=None):
    c = ws[ref]
    c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = align
    if border: c.border = border
    return c

def title(ws, ref, text):
    c = set_cell(ws, ref, text, font=F_TITLE, align=ALIGN_LEFT)
    return c

def section(ws, ref, text):
    c = set_cell(ws, ref, text, font=F_SECTION, fill=FILL_SECTION, align=ALIGN_LEFT)
    return c

def header(ws, ref, text):
    c = set_cell(ws, ref, text, font=F_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER)
    return c

def input_n(ws, ref, value, fmt=FMT_INT, highlight=False):
    f = F_INPUT_BOLD if highlight else F_INPUT
    fl = FILL_HIGHLIGHT if highlight else None
    return set_cell(ws, ref, value, font=f, fill=fl, fmt=fmt, align=ALIGN_RIGHT, border=BORDER)

def formula(ws, ref, expr, fmt=None, bold=False, fill=None):
    f = F_FORMULA_BOLD if bold else F_FORMULA
    return set_cell(ws, ref, expr, font=f, fill=fill, fmt=fmt, align=ALIGN_RIGHT, border=BORDER)

def label(ws, ref, text, bold=False):
    f = F_FORMULA_BOLD if bold else F_FORMULA
    return set_cell(ws, ref, text, font=f, align=ALIGN_LEFT, border=BORDER)

def note(ws, ref, text):
    return set_cell(ws, ref, text, font=F_NOTE, align=ALIGN_LEFT)

# =================================================================
# ABA 1 — PREMISSAS
# =================================================================
p = wb.create_sheet('Premissas')
p.column_dimensions['A'].width = 38
p.column_dimensions['B'].width = 16
p.column_dimensions['C'].width = 16
p.column_dimensions['D'].width = 50

title(p, 'A1', 'RR Engine — Premissas (v1)')
note(p, 'A2', 'Celulas em azul sao editaveis. Atualize aqui que o resto da planilha recalcula.')

# --- Configuracoes gerais ---
section(p, 'A4', 'Configuracoes gerais')
p.merge_cells('A4:D4')

label(p, 'A5', 'FX USD/BRL')
input_n(p, 'B5', 5.20, fmt='0.0000', highlight=True)
note(p, 'D5', 'Consistente com FX_USD_BRL em server/_core/llm-pricing.ts')

label(p, 'A6', 'Plano avulso (R$/orcamento)')
input_n(p, 'B6', 89.90, fmt=FMT_BRL)
note(p, 'D6', 'Pricing publicado em server/stripe/products.ts')

label(p, 'A7', 'Plano mensal (R$/mes)')
input_n(p, 'B7', 450.00, fmt=FMT_BRL)
label(p, 'A8', 'Quota mensal (orcamentos)')
input_n(p, 'B8', 10, fmt=FMT_INT)
label(p, 'A9', 'Receita media por orcamento no mensal (R$)', bold=True)
formula(p, 'B9', '=B7/B8', fmt=FMT_BRL, bold=True, fill=FILL_TOTAL)

label(p, 'A10', 'Stripe fee variavel (%)')
input_n(p, 'B10', 0.0399, fmt=FMT_PCT)
label(p, 'A11', 'Stripe fee fixo (R$)')
input_n(p, 'B11', 0.39, fmt=FMT_BRL)

# --- Custo de infraestrutura ---
section(p, 'A13', 'Custo de infraestrutura mensal')
p.merge_cells('A13:D13')

label(p, 'A14', 'Manus atual (R$/mes)')
input_n(p, 'B14', 0, fmt=FMT_BRL, highlight=True)
note(p, 'D14', 'Creditos legados — custo zero hoje')

label(p, 'A15', 'Pos-migracao otimista (R$/mes)')
input_n(p, 'B15', 550, fmt=FMT_BRL, highlight=True)
note(p, 'D15', 'Faixa otimista: R$ 400-700 (Vercel/Railway + TiDB + S3 + Anthropic direto). Media 550.')

# --- Volumes ---
section(p, 'A17', 'Volumes mensais por cenario de mercado')
p.merge_cells('A17:D17')

label(p, 'A18', 'Conservador (orcamentos/mes)')
input_n(p, 'B18', 10, fmt=FMT_INT)
note(p, 'D18', 'Validacao inicial — ~1 cliente regular')

label(p, 'A19', 'Moderado (orcamentos/mes)')
input_n(p, 'B19', 100, fmt=FMT_INT)
note(p, 'D19', '~3-4/dia util — 1 founder + 1 vendedor full-time')

label(p, 'A20', 'Otimista — ambicao 18 meses (orcamentos/mes)')
input_n(p, 'B20', 1000, fmt=FMT_INT)
note(p, 'D20', '~33/dia — exige forca de venda multiplicada')

# --- Tabela de precos LLM ---
section(p, 'A22', 'Precos por modelo (USD por milhao de tokens)')
p.merge_cells('A22:D22')

header(p, 'A23', 'Modelo')
header(p, 'B23', 'Input/Mtok')
header(p, 'C23', 'Output/Mtok')
set_cell(p, 'D23', 'Fonte', font=F_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT, border=BORDER)

modelos = [
    ('claude-opus-4-6', 15.00, 75.00, 'Anthropic API list price'),
    ('claude-sonnet-4-6', 3.00, 15.00, 'Anthropic API list price'),
    ('claude-haiku-4-5', 1.00, 5.00, 'Anthropic API list price'),
    ('gemini-2.5-flash', 0.30, 2.50, 'Google AI Studio list price'),
]
for i, (m, pin, pout, src) in enumerate(modelos):
    r = 24 + i
    label(p, f'A{r}', m)
    input_n(p, f'B{r}', pin, fmt=FMT_USD)
    input_n(p, f'C{r}', pout, fmt=FMT_USD)
    set_cell(p, f'D{r}', src, font=F_NOTE, align=ALIGN_LEFT, border=BORDER)

# --- Cenarios de tamanho de memorial ---
section(p, 'A30', 'Cenarios de tamanho de obra')
p.merge_cells('A30:D30')

header(p, 'A31', 'Cenario')
header(p, 'B31', 'Memorial (tokens)')
header(p, 'C31', 'Itens orcados')
set_cell(p, 'D31', 'Caso real de referencia', font=F_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT, border=BORDER)

label(p, 'A32', 'Pequeno')
input_n(p, 'B32', 5000, fmt=FMT_TOK)
input_n(p, 'C32', 30, fmt=FMT_INT)
set_cell(p, 'D32', 'Reforma residencial pequena (banheiro/cozinha)', font=F_NOTE, align=ALIGN_LEFT, border=BORDER)

label(p, 'A33', 'Medio')
input_n(p, 'B33', 15000, fmt=FMT_TOK)
input_n(p, 'C33', 80, fmt=FMT_INT)
set_cell(p, 'D33', 'Reforma Vania REV08', font=F_NOTE, align=ALIGN_LEFT, border=BORDER)

label(p, 'A34', 'Grande')
input_n(p, 'B34', 40000, fmt=FMT_TOK)
input_n(p, 'C34', 250, fmt=FMT_INT)
set_cell(p, 'D34', 'Hangar / obra industrial', font=F_NOTE, align=ALIGN_LEFT, border=BORDER)

# --- Bases ---
section(p, 'A36', 'Bases SINAPI/PINI')
p.merge_cells('A36:D36')
label(p, 'A37', 'Custo por consulta (R$)')
input_n(p, 'B37', 0, fmt=FMT_BRL)
note(p, 'D37', 'Base estatica + scraping orcamentor.com (gratis). PINI scraping pendente decisao P1.4.1.')

# =================================================================
# ABA 2 — CALCULO
# =================================================================
c = wb.create_sheet('Calculo')
for col, w in zip('ABCDEFGHIJKL', [28, 22, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]):
    c.column_dimensions[col].width = w

title(c, 'A1', 'Custo por orcamento (3 cenarios) — apos P0+P1')
note(c, 'A2', 'Tokens estimados a partir dos prompts em server/agents/index.ts. Modelos refletem P1.1 e P1.2.')

# Bloco de tokens por agente
section(c, 'A4', 'Estimativa de tokens por agente e cenario')
c.merge_cells('A4:L4')

# Cabecalho duplo
set_cell(c, 'A5', 'Agente', font=F_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER)
c.merge_cells('A5:A6')
set_cell(c, 'B5', 'Modelo', font=F_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER)
c.merge_cells('B5:B6')

set_cell(c, 'C5', 'Pequeno', font=F_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER)
c.merge_cells('C5:E5')
header(c, 'C6', 'Input')
header(c, 'D6', 'Output')
header(c, 'E6', 'Custo USD')

set_cell(c, 'F5', 'Medio', font=F_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER)
c.merge_cells('F5:H5')
header(c, 'F6', 'Input')
header(c, 'G6', 'Output')
header(c, 'H6', 'Custo USD')

set_cell(c, 'I5', 'Grande', font=F_HEADER, fill=FILL_HEADER, align=ALIGN_CENTER, border=BORDER)
c.merge_cells('I5:K5')
header(c, 'I6', 'Input')
header(c, 'J6', 'Output')
header(c, 'K6', 'Custo USD')

set_cell(c, 'L5', 'Notas', font=F_HEADER, fill=FILL_HEADER, align=ALIGN_LEFT, border=BORDER)
c.merge_cells('L5:L6')

# Linhas de agentes — modelo + tokens estimados
# (modelo, agente, in_p, out_p, in_m, out_m, in_g, out_g, nota)
agentes = [
    ('claude-opus-4-6', 'Engenheiro Tecnico', 9200, 6000, 19700, 14000, 45700, 30000, 'System prompt grande (~3.7k tokens) com tabelas de inferencia.'),
    ('gemini-2.5-flash', 'Logistica', 3900, 600, 7900, 1200, 16400, 2500, 'Default Gemini Flash; baixo custo por token.'),
    ('claude-opus-4-6', 'Orcamentista', 5900, 7000, 15900, 22000, 40900, 50000, 'Output dominante; price anchors injetadas top-15.'),
    ('claude-sonnet-4-6', 'Tributario', 3100, 800, 7600, 2000, 20600, 5500, 'Migrado para Sonnet em P1.1.'),
    ('claude-sonnet-4-6', 'Comercial', 0, 0, 0, 0, 0, 0, 'Deterministico desde P1.2 — zero custo LLM.'),
    ('claude-sonnet-4-6', 'Gestao Projetos', 4250, 5000, 8750, 14000, 19250, 28000, 'Cronograma diario detalhado. Sumarizado pos-P0.4.'),
    ('claude-sonnet-4-6', 'Financeiro', 0, 0, 0, 0, 0, 0, 'Deterministico desde P1.2 — zero custo LLM.'),
    ('claude-sonnet-4-6', 'Juridico', 550, 3000, 750, 5000, 1050, 7000, 'Migrado para Sonnet em P1.1; templating em P1.6 reduz output.'),
    ('claude-sonnet-4-6', 'Board', 3500, 2500, 5000, 3500, 8000, 5000, 'Migrado para Sonnet em P1.1.'),
    ('claude-sonnet-4-6', 'Auditor', 4500, 2000, 10000, 3500, 15000, 5500, 'Default Sonnet. Chunking pos-P0.4 pode multiplicar em obras grandes.'),
]

# Mapa: modelo -> linha em Premissas
linha_premissa = {
    'claude-opus-4-6': 24,
    'claude-sonnet-4-6': 25,
    'claude-haiku-4-5': 26,
    'gemini-2.5-flash': 27,
}

start = 7
for i, (modelo, agente, ip, op, im, om, ig, og, nota) in enumerate(agentes):
    r = start + i
    label(c, f'A{r}', agente)
    set_cell(c, f'B{r}', modelo, font=F_LINK, align=ALIGN_LEFT, border=BORDER)

    lp = linha_premissa[modelo]

    # Pequeno
    input_n(c, f'C{r}', ip, fmt=FMT_TOK)
    input_n(c, f'D{r}', op, fmt=FMT_TOK)
    formula(c, f'E{r}', f'=(C{r}*Premissas!B{lp}+D{r}*Premissas!C{lp})/1000000', fmt=FMT_USD)

    # Medio
    input_n(c, f'F{r}', im, fmt=FMT_TOK)
    input_n(c, f'G{r}', om, fmt=FMT_TOK)
    formula(c, f'H{r}', f'=(F{r}*Premissas!B{lp}+G{r}*Premissas!C{lp})/1000000', fmt=FMT_USD)

    # Grande
    input_n(c, f'I{r}', ig, fmt=FMT_TOK)
    input_n(c, f'J{r}', og, fmt=FMT_TOK)
    formula(c, f'K{r}', f'=(I{r}*Premissas!B{lp}+J{r}*Premissas!C{lp})/1000000', fmt=FMT_USD)

    set_cell(c, f'L{r}', nota, font=F_NOTE, align=ALIGN_LEFT, border=BORDER)

end = start + len(agentes) - 1

# Totais
tr = end + 1
label(c, f'A{tr}', 'Total tokens / Custo USD por orcamento', bold=True)
set_cell(c, f'B{tr}', '', fill=FILL_TOTAL, border=BORDER)

formula(c, f'C{tr}', f'=SUM(C{start}:C{end})', fmt=FMT_TOK, bold=True, fill=FILL_TOTAL)
formula(c, f'D{tr}', f'=SUM(D{start}:D{end})', fmt=FMT_TOK, bold=True, fill=FILL_TOTAL)
formula(c, f'E{tr}', f'=SUM(E{start}:E{end})', fmt=FMT_USD, bold=True, fill=FILL_TOTAL)
formula(c, f'F{tr}', f'=SUM(F{start}:F{end})', fmt=FMT_TOK, bold=True, fill=FILL_TOTAL)
formula(c, f'G{tr}', f'=SUM(G{start}:G{end})', fmt=FMT_TOK, bold=True, fill=FILL_TOTAL)
formula(c, f'H{tr}', f'=SUM(H{start}:H{end})', fmt=FMT_USD, bold=True, fill=FILL_TOTAL)
formula(c, f'I{tr}', f'=SUM(I{start}:I{end})', fmt=FMT_TOK, bold=True, fill=FILL_TOTAL)
formula(c, f'J{tr}', f'=SUM(J{start}:J{end})', fmt=FMT_TOK, bold=True, fill=FILL_TOTAL)
formula(c, f'K{tr}', f'=SUM(K{start}:K{end})', fmt=FMT_USD, bold=True, fill=FILL_TOTAL)

# Custo BRL
br = tr + 1
label(c, f'A{br}', 'Custo BRL por orcamento (USD * FX)', bold=True)
set_cell(c, f'B{br}', '', fill=FILL_HIGHLIGHT, border=BORDER)
for col, totcol in (('E', 'C'), ('H', 'F'), ('K', 'I')):
    pass

formula(c, f'E{br}', f'=E{tr}*Premissas!B5', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
formula(c, f'H{br}', f'=H{tr}*Premissas!B5', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
formula(c, f'K{br}', f'=K{tr}*Premissas!B5', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)

# Resumo no topo
section(c, 'A20', 'Resumo de custo por orcamento (BRL)')
c.merge_cells('A20:E20')

header(c, 'A21', 'Cenario')
header(c, 'B21', 'Tokens totais')
header(c, 'C21', 'Custo USD')
header(c, 'D21', 'Custo BRL')
header(c, 'E21', 'Modelos dominantes')

label(c, 'A22', 'Pequeno')
formula(c, 'B22', f'=C{tr}+D{tr}', fmt=FMT_TOK)
formula(c, 'C22', f'=E{tr}', fmt=FMT_USD)
formula(c, 'D22', f'=E{br}', fmt=FMT_BRL, bold=True)
set_cell(c, 'E22', 'Engenheiro+Orcamentista (Opus)', font=F_NOTE, align=ALIGN_LEFT, border=BORDER)

label(c, 'A23', 'Medio (Reforma Vania)')
formula(c, 'B23', f'=F{tr}+G{tr}', fmt=FMT_TOK)
formula(c, 'C23', f'=H{tr}', fmt=FMT_USD)
formula(c, 'D23', f'=H{br}', fmt=FMT_BRL, bold=True)
set_cell(c, 'E23', 'Engenheiro+Orcamentista (Opus)', font=F_NOTE, align=ALIGN_LEFT, border=BORDER)

label(c, 'A24', 'Grande (Hangar)')
formula(c, 'B24', f'=I{tr}+J{tr}', fmt=FMT_TOK)
formula(c, 'C24', f'=K{tr}', fmt=FMT_USD)
formula(c, 'D24', f'=K{br}', fmt=FMT_BRL, bold=True)
set_cell(c, 'E24', 'Engenheiro+Orcamentista (Opus)', font=F_NOTE, align=ALIGN_LEFT, border=BORDER)

# =================================================================
# ABA 3 — CENARIOS
# =================================================================
sc = wb.create_sheet('Cenarios')
for col, w in zip('ABCDEFGHIJ', [32, 16, 16, 16, 16, 16, 16, 16, 16, 16]):
    sc.column_dimensions[col].width = w

title(sc, 'A1', 'Cenarios de margem por plano e volume')
note(sc, 'A2', 'Plano avulso vs Plano mensal x 3 tamanhos x 3 volumes. Manus = 0 hoje. Pos-migracao usa Premissas!B15.')

# --- BLOCO PLANO AVULSO ---
section(sc, 'A4', 'Plano avulso (R$ 89,90 por orcamento)')
sc.merge_cells('A4:J4')

header(sc, 'A5', 'Item')
header(sc, 'B5', 'Pequeno')
header(sc, 'C5', 'Medio')
header(sc, 'D5', 'Grande')

label(sc, 'A6', 'Receita bruta')
formula(sc, 'B6', '=Premissas!B6', fmt=FMT_BRL)
formula(sc, 'C6', '=Premissas!B6', fmt=FMT_BRL)
formula(sc, 'D6', '=Premissas!B6', fmt=FMT_BRL)

label(sc, 'A7', 'Stripe fee (variavel + fixo)')
formula(sc, 'B7', '=B6*Premissas!B10+Premissas!B11', fmt=FMT_BRL)
formula(sc, 'C7', '=C6*Premissas!B10+Premissas!B11', fmt=FMT_BRL)
formula(sc, 'D7', '=D6*Premissas!B10+Premissas!B11', fmt=FMT_BRL)

label(sc, 'A8', 'Receita liquida', bold=True)
formula(sc, 'B8', '=B6-B7', fmt=FMT_BRL, bold=True, fill=FILL_TOTAL)
formula(sc, 'C8', '=C6-C7', fmt=FMT_BRL, bold=True, fill=FILL_TOTAL)
formula(sc, 'D8', '=D6-D7', fmt=FMT_BRL, bold=True, fill=FILL_TOTAL)

label(sc, 'A9', 'Custo LLM por orcamento')
formula(sc, 'B9', '=Calculo!D22', fmt=FMT_BRL)
formula(sc, 'C9', '=Calculo!D23', fmt=FMT_BRL)
formula(sc, 'D9', '=Calculo!D24', fmt=FMT_BRL)

label(sc, 'A10', 'Margem bruta (sem infra)', bold=True)
formula(sc, 'B10', '=B8-B9', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
formula(sc, 'C10', '=C8-C9', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
formula(sc, 'D10', '=D8-D9', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)

label(sc, 'A11', 'Margem bruta %')
formula(sc, 'B11', '=B10/B6', fmt=FMT_PCT)
formula(sc, 'C11', '=C10/C6', fmt=FMT_PCT)
formula(sc, 'D11', '=D10/D6', fmt=FMT_PCT)

# --- BLOCO PLANO MENSAL ---
section(sc, 'A13', 'Plano mensal (R$ 450/mes para 10 orcamentos = R$ 45/orcamento)')
sc.merge_cells('A13:J13')

header(sc, 'A14', 'Item')
header(sc, 'B14', 'Pequeno')
header(sc, 'C14', 'Medio')
header(sc, 'D14', 'Grande')

label(sc, 'A15', 'Receita bruta por orcamento')
formula(sc, 'B15', '=Premissas!B9', fmt=FMT_BRL)
formula(sc, 'C15', '=Premissas!B9', fmt=FMT_BRL)
formula(sc, 'D15', '=Premissas!B9', fmt=FMT_BRL)

label(sc, 'A16', 'Stripe fee alocado por orcamento')
formula(sc, 'B16', '=(Premissas!B7*Premissas!B10+Premissas!B11)/Premissas!B8', fmt=FMT_BRL)
formula(sc, 'C16', '=(Premissas!B7*Premissas!B10+Premissas!B11)/Premissas!B8', fmt=FMT_BRL)
formula(sc, 'D16', '=(Premissas!B7*Premissas!B10+Premissas!B11)/Premissas!B8', fmt=FMT_BRL)

label(sc, 'A17', 'Receita liquida por orcamento', bold=True)
formula(sc, 'B17', '=B15-B16', fmt=FMT_BRL, bold=True, fill=FILL_TOTAL)
formula(sc, 'C17', '=C15-C16', fmt=FMT_BRL, bold=True, fill=FILL_TOTAL)
formula(sc, 'D17', '=D15-D16', fmt=FMT_BRL, bold=True, fill=FILL_TOTAL)

label(sc, 'A18', 'Custo LLM por orcamento')
formula(sc, 'B18', '=Calculo!D22', fmt=FMT_BRL)
formula(sc, 'C18', '=Calculo!D23', fmt=FMT_BRL)
formula(sc, 'D18', '=Calculo!D24', fmt=FMT_BRL)

label(sc, 'A19', 'Margem bruta (sem infra)', bold=True)
formula(sc, 'B19', '=B17-B18', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
formula(sc, 'C19', '=C17-C18', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
formula(sc, 'D19', '=D17-D18', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)

label(sc, 'A20', 'Margem bruta %')
formula(sc, 'B20', '=B19/B15', fmt=FMT_PCT)
formula(sc, 'C20', '=C19/C15', fmt=FMT_PCT)
formula(sc, 'D20', '=D19/D15', fmt=FMT_PCT)

# --- BLOCO MARGEM CONSIDERANDO INFRA ---
section(sc, 'A22', 'Margem mensal consolidada — Manus = 0 hoje (Premissas!B14)')
sc.merge_cells('A22:J22')

header(sc, 'A23', 'Cenario')
header(sc, 'B23', 'Volume/mes')
header(sc, 'C23', 'Plano')
header(sc, 'D23', 'Receita liq mensal')
header(sc, 'E23', 'Custo LLM mensal')
header(sc, 'F23', 'Custo infra')
header(sc, 'G23', 'Margem mensal R$')
header(sc, 'H23', 'Margem %')

# Combinacoes: 3 volumes x 2 planos x 3 tamanhos = 18 linhas; foco em medio para clareza
combos = [
    # (label, volume_ref, plano, tamanho_col_calculo)
    ('Conservador / Avulso / Pequeno', 'Premissas!B18', 'avulso', 'B'),  # B = Pequeno em Cenarios
    ('Conservador / Avulso / Medio', 'Premissas!B18', 'avulso', 'C'),
    ('Conservador / Avulso / Grande', 'Premissas!B18', 'avulso', 'D'),
    ('Conservador / Mensal / Medio', 'Premissas!B18', 'mensal', 'C'),
    ('Moderado / Avulso / Medio', 'Premissas!B19', 'avulso', 'C'),
    ('Moderado / Mensal / Medio', 'Premissas!B19', 'mensal', 'C'),
    ('Otimista / Avulso / Medio', 'Premissas!B20', 'avulso', 'C'),
    ('Otimista / Mensal / Medio', 'Premissas!B20', 'mensal', 'C'),
]

linha_recliq = {'avulso': 8, 'mensal': 17}  # linha de receita liquida em Cenarios
linha_custollm = {'avulso': 9, 'mensal': 18}  # linha de custo LLM em Cenarios

start_combo = 24
for i, (lab, vol_ref, plano, tcol) in enumerate(combos):
    r = start_combo + i
    label(sc, f'A{r}', lab)
    formula(sc, f'B{r}', f'={vol_ref}', fmt=FMT_INT)
    set_cell(sc, f'C{r}', plano, font=F_LINK, align=ALIGN_LEFT, border=BORDER)
    formula(sc, f'D{r}', f'={tcol}{linha_recliq[plano]}*{vol_ref}', fmt=FMT_BRL)
    formula(sc, f'E{r}', f'={tcol}{linha_custollm[plano]}*{vol_ref}', fmt=FMT_BRL)
    formula(sc, f'F{r}', f'=Premissas!B14', fmt=FMT_BRL)  # Manus = 0
    formula(sc, f'G{r}', f'=D{r}-E{r}-F{r}', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
    formula(sc, f'H{r}', f'=IFERROR(G{r}/D{r},0)', fmt=FMT_PCT)

# Bloco pos-migracao
end_combo = start_combo + len(combos) - 1
section(sc, f'A{end_combo+2}', 'Mesma analise pos-migracao (infra = Premissas!B15)')
sc.merge_cells(f'A{end_combo+2}:J{end_combo+2}')

header(sc, f'A{end_combo+3}', 'Cenario')
header(sc, f'B{end_combo+3}', 'Volume/mes')
header(sc, f'C{end_combo+3}', 'Plano')
header(sc, f'D{end_combo+3}', 'Receita liq mensal')
header(sc, f'E{end_combo+3}', 'Custo LLM mensal')
header(sc, f'F{end_combo+3}', 'Custo infra')
header(sc, f'G{end_combo+3}', 'Margem mensal R$')
header(sc, f'H{end_combo+3}', 'Margem %')

start2 = end_combo + 4
for i, (lab, vol_ref, plano, tcol) in enumerate(combos):
    r = start2 + i
    label(sc, f'A{r}', lab)
    formula(sc, f'B{r}', f'={vol_ref}', fmt=FMT_INT)
    set_cell(sc, f'C{r}', plano, font=F_LINK, align=ALIGN_LEFT, border=BORDER)
    formula(sc, f'D{r}', f'={tcol}{linha_recliq[plano]}*{vol_ref}', fmt=FMT_BRL)
    formula(sc, f'E{r}', f'={tcol}{linha_custollm[plano]}*{vol_ref}', fmt=FMT_BRL)
    formula(sc, f'F{r}', f'=Premissas!B15', fmt=FMT_BRL)  # pos-migracao
    formula(sc, f'G{r}', f'=D{r}-E{r}-F{r}', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
    formula(sc, f'H{r}', f'=IFERROR(G{r}/D{r},0)', fmt=FMT_PCT)

# =================================================================
# ABA 4 — STRESS TEST
# =================================================================
st = wb.create_sheet('Stress Test')
for col, w in zip('ABCDEFGHIJ', [40, 16, 16, 16, 16, 16, 16, 16, 16, 16]):
    st.column_dimensions[col].width = w

title(st, 'A1', 'Stress test — sensibilidade a choques')
note(st, 'A2', 'Cenario base: Medio (Reforma Vania), plano avulso. Multiplicadores ajustaveis em azul.')

# --- Multiplicadores ---
section(st, 'A4', 'Multiplicadores de stress (1.0 = base, 2.0 = dobra)')
st.merge_cells('A4:D4')

label(st, 'A5', 'Multiplicador de tokens (memorial cresce)')
input_n(st, 'B5', 1.5, fmt='0.00', highlight=True)
note(st, 'D5', 'Ex: 1.5 = memorial cresce 50% em itens e tamanho')

label(st, 'A6', 'Multiplicador de preco LLM (provider sobe preco)')
input_n(st, 'B6', 1.0, fmt='0.00', highlight=True)
note(st, 'D6', 'Ex: 2.0 = Anthropic dobra precos')

label(st, 'A7', 'Stripe fee variavel (override)')
input_n(st, 'B7', 0.05, fmt=FMT_PCT, highlight=True)
note(st, 'D7', 'Ex: 5% se trocar gateway')

# --- Linha base de referencia ---
section(st, 'A9', 'Linha base — Medio, avulso (referencia em Cenarios)')
st.merge_cells('A9:D9')
label(st, 'A10', 'Receita bruta')
formula(st, 'B10', '=Premissas!B6', fmt=FMT_BRL)
label(st, 'A11', 'Receita liquida (base)')
formula(st, 'B11', '=Cenarios!C8', fmt=FMT_BRL)
label(st, 'A12', 'Custo LLM (base)')
formula(st, 'B12', '=Cenarios!C9', fmt=FMT_BRL)
label(st, 'A13', 'Margem bruta (base)', bold=True)
formula(st, 'B13', '=Cenarios!C10', fmt=FMT_BRL, bold=True, fill=FILL_TOTAL)
label(st, 'A14', 'Margem % (base)')
formula(st, 'B14', '=Cenarios!C11', fmt=FMT_PCT)

# --- Linha apos stress ---
section(st, 'A16', 'Apos aplicacao dos multiplicadores')
st.merge_cells('A16:D16')

label(st, 'A17', 'Receita bruta')
formula(st, 'B17', '=B10', fmt=FMT_BRL)
label(st, 'A18', 'Stripe fee (com override)')
formula(st, 'B18', '=B17*B7+Premissas!B11', fmt=FMT_BRL)
label(st, 'A19', 'Receita liquida (com override Stripe)', bold=True)
formula(st, 'B19', '=B17-B18', fmt=FMT_BRL, bold=True)
label(st, 'A20', 'Custo LLM (com multiplicadores)', bold=True)
formula(st, 'B20', '=B12*B5*B6', fmt=FMT_BRL, bold=True)
label(st, 'A21', 'Margem bruta apos stress', bold=True)
formula(st, 'B21', '=B19-B20', fmt=FMT_BRL, bold=True, fill=FILL_HIGHLIGHT)
label(st, 'A22', 'Margem % apos stress')
formula(st, 'B22', '=B21/B17', fmt=FMT_PCT)
label(st, 'A23', 'Variacao da margem (R$)')
formula(st, 'B23', '=B21-B13', fmt=FMT_BRL)

# --- Tabela de sensibilidade: ponto onde margem fica negativa ---
section(st, 'A25', 'Quando a margem do avulso (Medio) fica negativa?')
st.merge_cells('A25:H25')

label(st, 'A26', 'Variando multiplicador de tokens (preco LLM = base)')

header(st, 'B27', '1.0x')
header(st, 'C27', '1.5x')
header(st, 'D27', '2.0x')
header(st, 'E27', '3.0x')
header(st, 'F27', '5.0x')
header(st, 'G27', '10x')
header(st, 'H27', '20x')

label(st, 'A28', 'Margem bruta (R$)')
multipliers = [1.0, 1.5, 2.0, 3.0, 5.0, 10.0, 20.0]
cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
for col, m in zip(cols, multipliers):
    formula(st, f'{col}28', f'=Cenarios!C8-Cenarios!C9*{m}', fmt=FMT_BRL,
            bold=True, fill=FILL_HIGHLIGHT)

label(st, 'A29', 'Margem %')
for col, m in zip(cols, multipliers):
    formula(st, f'{col}29', f'=(Cenarios!C8-Cenarios!C9*{m})/Premissas!B6', fmt=FMT_PCT)

label(st, 'A30', 'Status')
for col, m in zip(cols, multipliers):
    formula(st, f'{col}30',
            f'=IF((Cenarios!C8-Cenarios!C9*{m})>0,"OK","NEGATIVA")',
            bold=True)

# --- Mesma analise para grande/avulso ---
section(st, 'A32', 'Mesma analise para Grande / avulso')
st.merge_cells('A32:H32')
header(st, 'B33', '1.0x')
header(st, 'C33', '1.5x')
header(st, 'D33', '2.0x')
header(st, 'E33', '3.0x')
header(st, 'F33', '5.0x')
header(st, 'G33', '10x')
header(st, 'H33', '20x')

label(st, 'A34', 'Margem bruta (R$)')
for col, m in zip(cols, multipliers):
    formula(st, f'{col}34', f'=Cenarios!D8-Cenarios!D9*{m}', fmt=FMT_BRL,
            bold=True, fill=FILL_HIGHLIGHT)
label(st, 'A35', 'Status')
for col, m in zip(cols, multipliers):
    formula(st, f'{col}35',
            f'=IF((Cenarios!D8-Cenarios!D9*{m})>0,"OK","NEGATIVA")',
            bold=True)

# --- Notas finais ---
section(st, 'A37', 'Observacoes')
st.merge_cells('A37:J37')
note(st, 'A38', '1. Numeros baseados em estimativas dos prompts (Fase 2 da analise). Substituir por dados reais quando agent_llm_calls tiver dataset (>= 5 orcamentos rodados).')
note(st, 'A39', '2. Margem nao inclui custos de aquisicao (CAC), suporte humano, ou atrasos por revisao manual. Esses entram em modelo P&L mais completo.')
note(st, 'A40', '3. Stripe Brasil cobra 3,99% + R$ 0,39 em cartao domestico; Pix tem fee menor. Revisar quando trocar gateway.')
note(st, 'A41', '4. Custo de bases SINAPI/PINI hoje e zero (scraping orcamentor.com gratuito). Se PINI scraping pago for ativado (P1.4.1), adicionar linha em Premissas e propagar.')

# =================================================================
# Salvar
# =================================================================
wb.save(OUT)
print(f'Gerado: {OUT}')
